import sys
import openpyxl
import time
import subprocess
import os
import urllib.request
import json
import websocket

_BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(__file__)
EXCEL_PATH = os.path.join(_BASE_DIR, "ivi.xlsx")
CHROME_PATH = r"C:\Program Files (x86)\Google\Chrome\chrome.exe"
DEBUG_PORT = 9222


def ensure_chrome_debug():
    """Проверяет порт 9222, если нет — перезапускает Chrome с remote debugging"""
    try:
        urllib.request.urlopen(f"http://localhost:{DEBUG_PORT}/json/version", timeout=2)
        return  # уже запущен
    except Exception:
        pass

    print("Starting Chrome with remote debugging...")
    os.system("taskkill /f /im chrome.exe 2>nul")
    time.sleep(2)

    subprocess.Popen(
        [CHROME_PATH, f"--remote-debugging-port={DEBUG_PORT}"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )

    for i in range(15):
        time.sleep(1)
        try:
            urllib.request.urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=2)
            print("Chrome ready!")
            return
        except Exception:
            print(".", end="", flush=True)
    raise RuntimeError("Chrome did not start with remote debugging")


def get_active_tab():
    r = urllib.request.urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=5)
    tabs = json.loads(r.read())
    for t in tabs:
        url = t.get("url", "")
        if "ivi" in url.lower() and "content" in url.lower():
            return t
    for t in tabs:
        url = t.get("url", "")
        if "ivi" in url.lower():
            return t
    return tabs[0] if tabs else None


def execute_js(tab, js):
    ws_url = tab["webSocketDebuggerUrl"]
    ws = websocket.create_connection(ws_url, timeout=10)
    msg = json.dumps({"id": 1, "method": "Runtime.evaluate", "params": {
        "expression": js, "returnByValue": True
    }})
    ws.send(msg)
    resp = json.loads(ws.recv())
    ws.close()
    return resp.get("result", {}).get("result", {}).get("value")


def read_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    midrolls = []
    for row in range(13, 19):
        val = ws.cell(row=row, column=9).value
        if val and val > 0:
            midrolls.append(int(val))

    return {
        "midrolls": midrolls,
        "start_scale": int(ws.cell(row=4, column=4).value or 0),
        "finish_scale": int(ws.cell(row=5, column=4).value or 0),
        "postroll": int(ws.cell(row=6, column=4).value or 0),
        "duration": int(ws.cell(row=7, column=4).value or 0),
    }


def fill_midrolls(tab, midrolls):
    if not midrolls:
        return
    print(f"  Filling {len(midrolls)} midrolls: {midrolls}")

    execute_js(tab, """
        // Удаляем существующие мидроллы
        document.querySelectorAll('input[name$="-DELETE"]').forEach(cb => {
            if (!cb.checked) cb.checked = true;
        });
    """)
    time.sleep(0.3)

    for i, val in enumerate(midrolls):
        if i > 0:
            execute_js(tab, """
                let btn = document.querySelector('.middroll_block a[data-action="add-formset"], .middroll_block a.add_link');
                if (!btn) btn = document.querySelector('a[data-action="add-formset"]');
                if (btn) btn.click();
            """)
            time.sleep(0.3)

        execute_js(tab, f"""
            let inp = document.getElementById('id_middroll-{i}-time');
            if (inp) {{ inp.value = '{val}'; inp.dispatchEvent(new Event('input', {{bubbles: true}})); }}
        """)


def fill_other(tab, data):
    print(f"  Start: {data['start_scale']}, Finish: {data['finish_scale']}")
    print(f"  Postroll: {data['postroll']}, Duration: {data['duration']}")

    execute_js(tab, f"""
        // Заставка
        let sel = document.querySelector('select[name$="-marker_type"]');
        if (sel) {{
            sel.value = '2';
            sel.dispatchEvent(new Event('change', {{bubbles: true}}));
            let row = sel.closest('tr');
            row.querySelector('input[name$="-start"]').value = '{data["start_scale"]}';
            row.querySelector('input[name$="-finish"]').value = '{data["finish_scale"]}';
        }}

        // Построл
        let cr = document.getElementById('id_credits_begin_time');
        if (cr) {{ cr.value = '{data["postroll"]}'; cr.dispatchEvent(new Event('input', {{bubbles: true}})); }}

        // Хронометраж
        let dur = document.getElementById('id_duration');
        if (dur) {{ dur.value = '{data["duration"]}'; dur.dispatchEvent(new Event('input', {{bubbles: true}})); }}
    """)


def main():
    data = read_excel()
    print("Excel data:")
    for k, v in data.items():
        print(f"  {k}: {v}")

    ensure_chrome_debug()

    tab = get_active_tab()
    if not tab:
        print("No tabs found in Chrome")
        return

    url = tab.get("url", "")
    print(f"\nCurrent page: {tab['title']}")
    print(f"URL: {url}")

    if "ivi" not in url.lower():
        print("\nChrome opened. Navigate to the admin content card, then press Enter...")
        input()
        tab = get_active_tab()

    fill_midrolls(tab, data["midrolls"])
    fill_other(tab, data)

    print("\nDone! Review and save in admin panel.")
    input("Press Enter to exit...")


if __name__ == "__main__":
    main()
