import json
import os
import threading
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    exit(1)

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
EXCEL_FILE = os.path.join(DESKTOP, "ivi_timestamps.xlsx")
HOST = "localhost"
PORT = 8765


def append_to_excel(video_name, timestamp_sec, url=""):
    exists = os.path.exists(EXCEL_FILE)
    if exists:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 60
        ws.append(["Video Name", "Timestamp (sec)", "Timestamp (HH:MM:SS)", "Captured At", "URL"])

    hours = int(timestamp_sec // 3600)
    minutes = int((timestamp_sec % 3600) // 60)
    seconds = int(timestamp_sec % 60)
    time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([video_name, round(timestamp_sec, 1), time_str, date_str, url])
    wb.save(EXCEL_FILE)
    print(f"[{date_str}] Saved: {video_name} @ {time_str} ({timestamp_sec}s)")


class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_POST(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length)
        data = json.loads(body)

        video_name = data.get("video_name", "Unknown")
        timestamp = data.get("timestamp", 0)
        url = data.get("url", "")

        append_to_excel(video_name, timestamp, url)

        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps({"status": "ok"}).encode())

    def log_message(self, format, *args):
        pass


def run_server():
    server = HTTPServer((HOST, PORT), Handler)
    print(f"[IVI Timestamp Server] Running on http://{HOST}:{PORT}")
    print(f"[IVI Timestamp Server] Excel file: {EXCEL_FILE}")
    print("[IVI Timestamp Server] Press Ctrl+C to stop")
    server.serve_forever()


if __name__ == "__main__":
    run_server()
