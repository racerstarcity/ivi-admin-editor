import json
import sys
import openpyxl
import subprocess
import time
import os
import urllib.request
import websocket
from http.server import HTTPServer, BaseHTTPRequestHandler

_BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(__file__)
EXCEL_PATH = os.path.join(_BASE_DIR, "ivi.xlsx")
CHROME_PATH = r"C:\Program Files (x86)\Google\Chrome\chrome.exe"
DEBUG_PORT = 9222
DATA_PORT = 8766


def read_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    midrolls = []
    for row in range(13, 19):
        val = ws.cell(row=row, column=9).value
        if val and val > 0:
            midrolls.append(int(val))
    return {
        "midrolls": midrolls,
        "start_prev": int(ws.cell(row=2, column=4).value or 0),
        "finish_prev": int(ws.cell(row=3, column=4).value or 0),
        "start_scale": int(ws.cell(row=4, column=4).value or 0),
        "finish_scale": int(ws.cell(row=5, column=4).value or 0),
        "postroll": int(ws.cell(row=6, column=4).value or 0),
        "duration": int(ws.cell(row=7, column=4).value or 0),
    }


def js_code(data):
    return """
fetch('http://localhost:8766/api').then(r=>r.json()).then(function(d){
function a(e){
e.dispatchEvent(new KeyboardEvent('keydown',{key:'ArrowRight',keyCode:39,bubbles:true}));
e.dispatchEvent(new KeyboardEvent('keyup',{key:'ArrowRight',keyCode:39,bubbles:true}));}
function aa(e){a(e);a(e);}
for(var i=1;i<d.midrolls.length;i++){
var tf=document.getElementById('id_middroll-TOTAL_FORMS');
tf.value=parseInt(tf.value)+1;
var first=document.querySelector('.middroll_block .middroll_formset_tr');
var row=first.cloneNode(true);
row.querySelectorAll('[name],[id]').forEach(function(el){
if(el.name)el.name=el.name.replace('middroll-0-','middroll-'+i+'-');
if(el.id)el.id=el.id.replace('middroll-0-','middroll-'+i+'-');
if(el.type=='checkbox')el.checked=false;
if(el.type=='text'||el.type=='hidden')el.value='';});
first.parentNode.insertBefore(row,first.nextSibling);}
d.midrolls.forEach(function(v,i){
var inp=document.getElementById('id_middroll-'+i+'-time');
if(inp){inp.value=v;inp.dispatchEvent(new Event('input',{bubbles:true}));aa(inp);}});
if(d.finish_scale>0){
var s0=document.getElementById('id_localization_labels-0-marker_type');
if(s0){s0.value='2';s0.dispatchEvent(new Event('change',{bubbles:true}));
document.getElementById('id_localization_labels-0-start').value=d.start_scale;a(document.getElementById('id_localization_labels-0-start'));
document.getElementById('id_localization_labels-0-finish').value=d.finish_scale;a(document.getElementById('id_localization_labels-0-finish'));}}
if(d.finish_prev>0){
var tf2=document.querySelector('[name="localization_labels-TOTAL_FORMS"]');
var n=parseInt(tf2.value);tf2.value=n+1;
var tbody=document.querySelector('.localization_labels_tr').closest('tbody');
var first2=tbody.querySelector('.localization_labels_tr');
var row2=first2.cloneNode(true);
row2.querySelectorAll('[name],[id]').forEach(function(el){
if(el.name)el.name=el.name.replace('localization_labels-0-','localization_labels-'+n+'-');
if(el.id)el.id=el.id.replace('localization_labels-0-','localization_labels-'+n+'-');
if(el.type=='checkbox')el.checked=false;
if(el.type!='select-one')el.value='';});
tbody.appendChild(row2);
document.getElementById('id_localization_labels-'+n+'-marker_type').value='1';
document.getElementById('id_localization_labels-'+n+'-start').value=d.start_prev;a(document.getElementById('id_localization_labels-'+n+'-start'));
document.getElementById('id_localization_labels-'+n+'-finish').value=d.finish_prev;a(document.getElementById('id_localization_labels-'+n+'-finish'));}
document.getElementById('id_credits_begin_time').value=d.postroll;a(document.getElementById('id_credits_begin_time'));
document.getElementById('id_duration').value=d.duration;a(document.getElementById('id_duration'));
document.querySelectorAll('.localization_formset_tr').forEach(function(row){
var lang=row.querySelector('select[name$="-localization_type"]');
if(!lang||!lang.value)return;
var idx=lang.name.match(/localizations-(\\\\d+)-/);
if(!idx)return;
var i=idx[1];
var dur=row.querySelector('[name="localizations-'+i+'-duration"]');
if(dur){dur.value=d.duration;a(dur);}
var cr=row.querySelector('[name="localizations-'+i+'-credits_begin_time"]');
if(cr){cr.value=d.postroll;a(cr);}
if(d.finish_scale>0){
var m0=row.querySelector('[name="custom_localization_labels-localizations-'+i+'-form-0-marker_type"]');
if(m0){m0.value='2';m0.dispatchEvent(new Event('change',{bubbles:true}));
var ms0=row.querySelector('[name="custom_localization_labels-localizations-'+i+'-form-0-start"]');
if(ms0){ms0.value=d.start_scale;a(ms0);}
var mf0=row.querySelector('[name="custom_localization_labels-localizations-'+i+'-form-0-finish"]');
if(mf0){mf0.value=d.finish_scale;a(mf0);}}}
if(d.finish_prev>0){
var addBtn=row.querySelector('[data-action="add-inner-form"]');
if(addBtn)addBtn.click();
var tid='id_custom_localization_labels-localizations-'+i+'-form-TOTAL_FORMS';
var tf3=document.getElementById(tid);
if(tf3)tf3.value=parseInt(tf3.value)+1;
var nestTbody=row.querySelector('.localization_markers_table tbody');
var emptyForm=nestTbody.querySelector('.empty-form');
var newRow=emptyForm.cloneNode(true);
newRow.className='nested-loc-form';
newRow.classList.remove('empty-form');
newRow.querySelectorAll('[name],[id]').forEach(function(el){
if(el.name)el.name=el.name.replace('__prefix__','1');
if(el.id)el.id=el.id.replace('__prefix__','1');
if(el.type=='checkbox')el.checked=false;
if(el.type!='select-one')el.value='';});
emptyForm.parentNode.insertBefore(newRow,emptyForm);
var m1=newRow.querySelector('[name="custom_localization_labels-localizations-'+i+'-form-1-marker_type"]');
if(m1){m1.value='1';m1.dispatchEvent(new Event('change',{bubbles:true}));
var ms1=newRow.querySelector('[name="custom_localization_labels-localizations-'+i+'-form-1-start"]');
if(ms1){ms1.value=d.start_prev;a(ms1);}
var mf1=newRow.querySelector('[name="custom_localization_labels-localizations-'+i+'-form-1-finish"]');
if(mf1){mf1.value=d.finish_prev;a(mf1);}}}});
alert('Done! Filled from Excel');
}).catch(function(e){alert('Error: '+e.message)});
""".replace("\\\\d", "\\d")


def ensure_chrome():
    try:
        urllib.request.urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=2)
        return True
    except Exception:
        pass

    print("Starting Chrome with remote debugging...")
    os.system("taskkill /f /im chrome.exe 2>nul")
    time.sleep(2)
    subprocess.Popen(
        [CHROME_PATH, f"--remote-debugging-port={DEBUG_PORT}", "--no-first-run"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    for i in range(30):
        time.sleep(1)
        try:
            urllib.request.urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=2)
            print("Chrome ready!")
            return True
        except Exception:
            pass
    return False


def get_active_tab():
    r = urllib.request.urlopen(f"http://localhost:{DEBUG_PORT}/json", timeout=5)
    tabs = json.loads(r.read())
    for t in tabs:
        url = t.get("url", "")
        if "ivi" in url.lower() and "content" in url.lower():
            return t
    for t in tabs:
        url = t.get("url", "")
        if "b2b.ivi" in url.lower():
            return t
    return tabs[0] if tabs else None


def execute_js(tab, js):
    ws_url = tab["webSocketDebuggerUrl"]
    ws = websocket.create_connection(ws_url, timeout=15)
    msg = json.dumps({"id": 1, "method": "Runtime.evaluate", "params": {
        "expression": js, "returnByValue": True
    }})
    ws.send(msg)
    resp = json.loads(ws.recv())
    ws.close()
    return resp


class DataHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/api":
            data = read_excel()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps(data).encode())
        else:
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.end_headers()
            self.wfile.write(b"IVI Fill Admin - Server running")
    def log_message(self, *args):
        pass


def start_data_server():
    server = HTTPServer(("localhost", DATA_PORT), DataHandler)
    import threading
    t = threading.Thread(target=server.serve_forever, daemon=True)
    t.start()
    return server


def main():
    print("=" * 50)
    print("  IVI Fill Admin")
    print("=" * 50)

    data = read_excel()
    print("\nExcel data:")
    for k, v in data.items():
        print(f"  {k}: {v}")
    print()

    if not ensure_chrome():
        input("Chrome did not start. Press Enter to exit...")
        return

    start_data_server()

    print("\nChrome is open. Navigate to the content card in the admin panel.")
    input("Press Enter when ready (on the content page)...")

    tab = get_active_tab()
    if not tab:
        print("No active tab found.")
        input("Press Enter to exit...")
        return

    print(f"\nPage: {tab['title']}")
    print(f"URL: {tab.get('url', '')}")

    print("\nInjecting data...")
    result = execute_js(tab, js_code(data))
    print("Done!")

    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
